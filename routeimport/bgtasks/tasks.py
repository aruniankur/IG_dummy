from celery import shared_task 
from time import sleep
from run import app
import pandas as pd
from models import db, User, Data, Workstation,Labor,ItemCustomField,ItemFinance,ItemFinance,ItemCategory,Category,Customer, ZohoInfo, UserDataMapping, Subscription, SubDataMapping, Company, DataConfiguration, Item, BOM
from openpyxl import Workbook
import openpyxl
import json


celery_app = app.extensions["celery"] 

@shared_task(ignore_result=False) 
def long_running_task(iterations):
    result = 0
    for i in range(iterations):
        result += i
        sleep(2) 
    return result 

@shared_task(ignore_result=False) 
def long_running_task2(iterations):
    result = 0
    for i in range(iterations):
        result += i
        sleep(2) 
    return result 

@shared_task(ignore_result=False)
def update_bom_linking(csv_file_path, data_id):
    database = Data.query.filter_by(id = data_id).first()
    df = pd.read_csv(csv_file_path)
    res = ""
    for index, row in df.iterrows():
        parent_item_code = str(row["PARENT CODE"])
        child_item_code = str(row["CHILD CODE"])
        child_item_qty = float(row["CHILD QTY"])
        margin = row['MARGIN']
        margin = margin if margin else 0
        parent_item = Item.query.filter_by(database = database, code = parent_item_code).first()
        child_item = Item.query.filter_by(database = database, code = child_item_code).first()
        if not parent_item or not child_item:
            res+= f"Error in Relation b/w {parent_item_code} and {child_item_code}. "
            print(f"Error in Relation b/w {parent_item_code} and {child_item_code}.")
            continue
        bom_item = BOM(database = database, parent_item = parent_item, child_item = child_item, child_item_qty=child_item_qty,
         child_item_unit = child_item.unit, margin=margin)
        db.session.add(bom_item)
        res += f"Added Relation b/w {parent_item.name} and {child_item.name}. "
    db.session.commit()
    return res

@shared_task(ignore_result=False)
def update_category_linking(csv_file_path, data_id):
    database = Data.query.filter_by(id = data_id).first()
    df = pd.read_csv(csv_file_path)
    print(df)
    # Assuming the first column is 'Name' and the rest are category names
    item_names = df['Item Name']
    category_names = df.columns[1:]
    category_dict = {}
    for category_name in category_names:
        cat_check = Category.query.filter_by(name=category_name, database=database).first()
        if cat_check:
            category_dict[category_name] = cat_check
    for index, item_name in enumerate(item_names):
        item = Item.query.filter_by(name=item_name, database=database).first()
        if item:
            for category_name in category_dict.keys():
                if df.at[index, category_name] == 1:
                    # Add the link if it doesn't exist
                    if not ItemCategory.query.filter_by(item=item, category=category_dict[category_name], database=database).first():
                        item_category = ItemCategory(item=item, category=category_dict[category_name], database=database)
                        db.session.add(item_category)
                else:
                    # Remove the link if it exists
                    existing_link = ItemCategory.query.filter_by(item=item, category=category_dict[category_name], database=database).first()
                    if existing_link:
                        db.session.delete(existing_link)
    db.session.commit()
    return True


@shared_task(ignore_result=False)
def partnerMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id = database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    res_string=""
    for i in range(1, sheet.max_row):
        name=sheet.cell(row=i+1,column=1).value
        partner_check = Customer.query.filter_by(name=name, database=database).first()
        if partner_check:
            res_string += "Excel contains already existing Partner names! Failed to add "+name+"\n"
            continue
        else:
            partner = Customer(name = sheet.cell(row=i+1, column=1).value,shipping_address=sheet.cell(row=i+1, column=3).value, billing_address=sheet.cell(row=i+1,column=2).value,
             gst=sheet.cell(row=i+1,column=4).value, email= sheet.cell(row=i+1, column=5).value, phone= sheet.cell(row=i+1, column=6).value,database=database)
            # print(partner.name, partner.code, partner.rate, partner.unit)
            db.session.add(partner)
            db.session.commit()
            res_string+=f"Added partner {name} successfully"
    return res_string


@shared_task(ignore_result=False)
def itemMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id=database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]  # Set column names from the first row
    df = df.iloc[1:]  # Remove the first row

    # Filter rows with empty names
    df = df[df['NAME'].notnull()]
    df = df[df['NAME'].str.strip() != '']
    res_string = ""

    # Split the dataframe into two parts
    # match_key = ['NAME', 'name']
    match_key = ['CODE', 'code']

    items_df = pd.DataFrame(db.session.query(Item.name, Item.code).filter(Item.data_id == database_id).all(), columns=["name", "code"])
    # print(items_df)
    new_items_df = df[~df[match_key[0]].isin(items_df[match_key[1]].tolist())]
    existing_items_df = df[df[match_key[0]].isin(items_df[match_key[1]].tolist())]
    print(existing_items_df, new_items_df)
    # Add new items

    data_config = DataConfiguration.query.filter_by(database=database).first()
    if not data_config:
        config_dict = {'ADDITIONAL_FIELDS':[], 'SEARCH_FIELDS':[]}
        new_data_config = DataConfiguration(database = database, item_master_config=json.dumps(config_dict))
        db.session.add(new_data_config)
        db.session.commit()
        data_config= new_data_config
    item_master_config_dict = json.loads(data_config.item_master_config)
    print(item_master_config_dict)
    for index, row in new_items_df.iterrows():
        item = Item(code=row.get('CODE', None), name=row.get('NAME', None), rate=row.get('JOB RATE', 0),
                    unit=row.get('UNIT', None), database=database)
        db.session.add(item)
        db.session.commit()

        item_finance = ItemFinance(database=database, item=item, cost_price=row.get('COST PRICE', 0),
                                   sale_price=row.get('SALE PRICE', 0), tax=row.get('TAX(%)', 0), hsn_code=row.get('HSN_CODE', None))
        db.session.add(item_finance)
        db.session.commit()

        additional_fields = ItemCustomField.query.filter_by(database=database, item=item).all()
        for field in item_master_config_dict["ADDITIONAL_FIELDS"]:
            # print(field)
            field_name = field["name"]
            additional_field_edit_value = row.get(field_name, None)
            if additional_field_edit_value:
                item_custom_field = ItemCustomField.query.filter_by(database=database, item=item, field_name=field_name).first()
                if not item_custom_field:
                    item_custom_field = ItemCustomField(field_name = field_name, field_value=additional_field_edit_value, item=item, database=database)
                    db.session.add(item_custom_field)
                    db.session.commit()
                else:
                    item_custom_field.field_value = additional_field_edit_value
                    db.session.commit()
            else:
                print(f"Field {field_name} not found!")
        res_string += f"Added item {row.get('NAME', None)} successfully\n"

    # Update existing items
    for index, row in existing_items_df.iterrows():
        item = Item.query.filter_by(code=row.get(match_key[0], None), database=database).first()
        if item:
            item.name = row.get('NAME', None)
            item.code = row.get('CODE', None)
            item.rate = row.get('JOB RATE', 0)
            item.unit = row.get('UNIT', None)

            item_finance = ItemFinance.query.filter_by(item=item).first()
            if item_finance:
                item_finance.cost_price = row.get('COST PRICE', 0)
                item_finance.sale_price = row.get('SALE PRICE', 0)
                item_finance.tax = row.get('TAX(%)', 0)
                item_finance.hsn_code = row.get('HSN_CODE', None)

            db.session.commit()

            additional_fields = ItemCustomField.query.filter_by(database=database, item=item).all()
            for field in item_master_config_dict["ADDITIONAL_FIELDS"]:
                # print(field)
                field_name = field["name"]
                additional_field_edit_value = row.get(field_name, None)
                if additional_field_edit_value:
                    item_custom_field = ItemCustomField.query.filter_by(database=database, item=item, field_name=field_name).first()
                    if not item_custom_field:
                        item_custom_field = ItemCustomField(field_name = field_name, field_value=additional_field_edit_value, item=item, database=database)
                        db.session.add(item_custom_field)
                        db.session.commit()
                    else:
                        item_custom_field.field_value = additional_field_edit_value
                        db.session.commit()
                else:
                    print(f"Field {field_name} not found!")
            res_string += f"Updated item {row.get('NAME', None)} successfully\n"
        else:
            res_string += f"Item {row.get('NAME', None)} not found in the database. Skipped updating.\n"

    return res_string


@shared_task(ignore_result=False)
def resourceMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id=database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]  # Set column names from the first row
    df = df.iloc[1:]  # Remove the first row

    # Filter rows with empty names
    df = df[df['NAME'].notnull()]
    df = df[df['NAME'].str.strip() != '']
    res_string = ""

    # Split the dataframe into two parts
    items_df = pd.DataFrame(db.session.query(Labor.name).filter(Labor.data_id == database_id).all(), columns=["name"])
    # print(items_df)
    new_items_df = df[~df['NAME'].isin(items_df['name'].tolist())]
    existing_items_df = df[df['NAME'].isin(items_df['name'].tolist())]
    print(existing_items_df, new_items_df)
    for index, row in new_items_df.iterrows():
        name=row.get('NAME', None)
        if name:
            item = Labor(code=row.get('CODE', "NA"), name=name, salary=row.get('HOURLY COST', 0),
                        gender=row.get('TYPE', "NA"), database=database)
            db.session.add(item)
            db.session.commit()
        res_string += f"Added item {row.get('NAME', None)} successfully\n"

    # Update existing items
    for index, row in existing_items_df.iterrows():
        name=row.get('NAME', None)
        if name:
            item = Labor.query.filter_by(name=name, database=database).first()
            if item:
                item.code = row.get('CODE', None)
                item.salary = row.get('HOURLY COST', 0)
                item.gender = row.get('UNIT', None)

                db.session.commit()
                res_string += f"Updated item {row.get('NAME', None)} successfully\n"
            else:
                res_string += f"Item {row.get('NAME', None)} not found in the database. Skipped updating.\n"

    return res_string


