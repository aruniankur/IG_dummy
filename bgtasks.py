from celery import shared_task 
from time import sleep
from config import app
import pandas as pd
from models import db, Data,ItemUnit,ItemInventory,Labor,ItemCustomField,ItemFinance,ItemFinance,ItemCategory,Category,Customer, DataConfiguration, Item, BOM
import openpyxl
import json



celery_app = app.extensions["celery"] 

@shared_task(ignore_result=False) 
def long_running_task(iterations):
    result = 0
    for i in range(iterations):
        result += i
        sleep(2) 
    return result 

@shared_task(ignore_result=False) 
def long_running_task2(iterations):
    result = 0
    for i in range(iterations):
        result += i
        sleep(2) 
    return result 

@shared_task(ignore_result=False)
def update_bom_linking(csv_file_path, data_id):
    database = Data.query.filter_by(id = data_id).first()
    df = pd.read_csv(csv_file_path)
    res = ""
    for index, row in df.iterrows():
        parent_item_code = str(row["PARENT CODE"])
        child_item_code = str(row["CHILD CODE"])
        child_item_qty = float(row["CHILD QTY"])
        margin = row['MARGIN']
        margin = margin if margin else 0
        parent_item = Item.query.filter_by(database = database, code = parent_item_code).first()
        child_item = Item.query.filter_by(database = database, code = child_item_code).first()
        if not parent_item or not child_item:
            res+= f"Error in Relation b/w {parent_item_code} and {child_item_code}. "
            print(f"Error in Relation b/w {parent_item_code} and {child_item_code}.")
            continue
        bom_item = BOM(database = database, parent_item = parent_item, child_item = child_item, child_item_qty=child_item_qty,
         child_item_unit = child_item.unit, margin=margin)
        db.session.add(bom_item)
        res += f"Added Relation b/w {parent_item.name} and {child_item.name}. "
    db.session.commit()
    return res

@shared_task(ignore_result=False)
def update_category_linking(csv_file_path, data_id):
    database = Data.query.filter_by(id = data_id).first()
    df = pd.read_csv(csv_file_path)
    print(df)
    # Assuming the first column is 'Name' and the rest are category names
    item_names = df['Item Name']
    category_names = df.columns[1:]
    category_dict = {}
    for category_name in category_names:
        cat_check = Category.query.filter_by(name=category_name, database=database).first()
        if cat_check:
            category_dict[category_name] = cat_check
    for index, item_name in enumerate(item_names):
        item = Item.query.filter_by(name=item_name, database=database).first()
        if item:
            for category_name in category_dict.keys():
                if df.at[index, category_name] == 1:
                    # Add the link if it doesn't exist
                    if not ItemCategory.query.filter_by(item=item, category=category_dict[category_name], database=database).first():
                        item_category = ItemCategory(item=item, category=category_dict[category_name], database=database)
                        db.session.add(item_category)
                else:
                    # Remove the link if it exists
                    existing_link = ItemCategory.query.filter_by(item=item, category=category_dict[category_name], database=database).first()
                    if existing_link:
                        db.session.delete(existing_link)
    db.session.commit()
    return True


@shared_task(ignore_result=False)
def partnerMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id = database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    res_string=""
    for i in range(1, sheet.max_row):
        name=sheet.cell(row=i+1,column=1).value
        partner_check = Customer.query.filter_by(name=name, database=database).first()
        if partner_check:
            res_string += "Excel contains already existing Partner names! Failed to add "+name+"\n"
            continue
        else:
            partner = Customer(name = sheet.cell(row=i+1, column=1).value,shipping_address=sheet.cell(row=i+1, column=3).value, billing_address=sheet.cell(row=i+1,column=2).value,
             gst=sheet.cell(row=i+1,column=4).value, email= sheet.cell(row=i+1, column=5).value, phone= sheet.cell(row=i+1, column=6).value,database=database)
            # print(partner.name, partner.code, partner.rate, partner.unit)
            db.session.add(partner)
            db.session.commit()
            res_string+=f"Added partner {name} successfully"
    return res_string


@shared_task(ignore_result=False)
def itemMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id=database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]  # Set column names from the first row
    df = df.iloc[1:]  # Remove the first row

    # Filter rows with empty names
    df = df[df['NAME'].notnull()]
    df = df[df['NAME'].str.strip() != '']
    res_string = ""

    # Split the dataframe into two parts
    # match_key = ['NAME', 'name']
    match_key = ['CODE', 'code']

    items_df = pd.DataFrame(db.session.query(Item.name, Item.code).filter(Item.data_id == database_id).all(), columns=["name", "code"])
    # print(items_df)
    new_items_df = df[~df[match_key[0]].isin(items_df[match_key[1]].tolist())]
    existing_items_df = df[df[match_key[0]].isin(items_df[match_key[1]].tolist())]
    print(existing_items_df, new_items_df)
    # Add new items

    data_config = DataConfiguration.query.filter_by(database=database).first()
    if not data_config:
        config_dict = {'ADDITIONAL_FIELDS':[], 'SEARCH_FIELDS':[]}
        new_data_config = DataConfiguration(database = database, item_master_config=json.dumps(config_dict))
        db.session.add(new_data_config)
        db.session.commit()
        data_config= new_data_config
    item_master_config_dict = json.loads(data_config.item_master_config)
    print(item_master_config_dict)
    for index, row in new_items_df.iterrows():
        item = Item(code=row.get('CODE', None), name=row.get('NAME', None), rate=row.get('JOB RATE', 0),
                    unit=row.get('UNIT', None), database=database)
        db.session.add(item)
        db.session.commit()

        item_finance = ItemFinance(database=database, item=item, cost_price=row.get('COST PRICE', 0),
                                   sale_price=row.get('SALE PRICE', 0), tax=row.get('TAX(%)', 0), hsn_code=row.get('HSN_CODE', None))
        db.session.add(item_finance)
        db.session.commit()

        additional_fields = ItemCustomField.query.filter_by(database=database, item=item).all()
        for field in item_master_config_dict["ADDITIONAL_FIELDS"]:
            # print(field)
            field_name = field["name"]
            additional_field_edit_value = row.get(field_name, None)
            if additional_field_edit_value:
                item_custom_field = ItemCustomField.query.filter_by(database=database, item=item, field_name=field_name).first()
                if not item_custom_field:
                    item_custom_field = ItemCustomField(field_name = field_name, field_value=additional_field_edit_value, item=item, database=database)
                    db.session.add(item_custom_field)
                    db.session.commit()
                else:
                    item_custom_field.field_value = additional_field_edit_value
                    db.session.commit()
            else:
                print(f"Field {field_name} not found!")
        res_string += f"Added item {row.get('NAME', None)} successfully\n"

    # Update existing items
    for index, row in existing_items_df.iterrows():
        item = Item.query.filter_by(code=row.get(match_key[0], None), database=database).first()
        if item:
            item.name = row.get('NAME', None)
            item.code = row.get('CODE', None)
            item.rate = row.get('JOB RATE', 0)
            item.unit = row.get('UNIT', None)

            item_finance = ItemFinance.query.filter_by(item=item).first()
            if item_finance:
                item_finance.cost_price = row.get('COST PRICE', 0)
                item_finance.sale_price = row.get('SALE PRICE', 0)
                item_finance.tax = row.get('TAX(%)', 0)
                item_finance.hsn_code = row.get('HSN_CODE', None)

            db.session.commit()

            additional_fields = ItemCustomField.query.filter_by(database=database, item=item).all()
            for field in item_master_config_dict["ADDITIONAL_FIELDS"]:
                # print(field)
                field_name = field["name"]
                additional_field_edit_value = row.get(field_name, None)
                if additional_field_edit_value:
                    item_custom_field = ItemCustomField.query.filter_by(database=database, item=item, field_name=field_name).first()
                    if not item_custom_field:
                        item_custom_field = ItemCustomField(field_name = field_name, field_value=additional_field_edit_value, item=item, database=database)
                        db.session.add(item_custom_field)
                        db.session.commit()
                    else:
                        item_custom_field.field_value = additional_field_edit_value
                        db.session.commit()
                else:
                    print(f"Field {field_name} not found!")
            res_string += f"Updated item {row.get('NAME', None)} successfully\n"
        else:
            res_string += f"Item {row.get('NAME', None)} not found in the database. Skipped updating.\n"

    return res_string


@shared_task(ignore_result=False)
def resourceMasterUpload(database_id, file_path):
    database = Data.query.filter_by(id=database_id).first()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]  # Set column names from the first row
    df = df.iloc[1:]  # Remove the first row

    # Filter rows with empty names
    df = df[df['NAME'].notnull()]
    df = df[df['NAME'].str.strip() != '']
    res_string = ""

    # Split the dataframe into two parts
    items_df = pd.DataFrame(db.session.query(Labor.name).filter(Labor.data_id == database_id).all(), columns=["name"])
    # print(items_df)
    new_items_df = df[~df['NAME'].isin(items_df['name'].tolist())]
    existing_items_df = df[df['NAME'].isin(items_df['name'].tolist())]
    print(existing_items_df, new_items_df)
    for index, row in new_items_df.iterrows():
        name=row.get('NAME', None)
        if name:
            item = Labor(code=row.get('CODE', "NA"), name=name, salary=row.get('HOURLY COST', 0),
                        gender=row.get('TYPE', "NA"), database=database)
            db.session.add(item)
            db.session.commit()
        res_string += f"Added item {row.get('NAME', None)} successfully\n"

    # Update existing items
    for index, row in existing_items_df.iterrows():
        name=row.get('NAME', None)
        if name:
            item = Labor.query.filter_by(name=name, database=database).first()
            if item:
                item.code = row.get('CODE', None)
                item.salary = row.get('HOURLY COST', 0)
                item.gender = row.get('UNIT', None)

                db.session.commit()
                res_string += f"Updated item {row.get('NAME', None)} successfully\n"
            else:
                res_string += f"Item {row.get('NAME', None)} not found in the database. Skipped updating.\n"

    return res_string


@shared_task(ignore_result=False)
def itemCopyData(database_id, checkboxes, from_item_id, to_item_ids):
    database = Data.query.filter_by(id = database_id).first()
    res = "Copied on "
    from_item = Item.query.filter_by(database=database, id=from_item_id).first()
    bom_items = BOM.query.filter_by(database=database, parent_item = from_item).all()
    item_categories = ItemCategory.query.filter_by(database=database, item=from_item).all()
    item_units = ItemUnit.query.filter_by(database=database, item=from_item).all()
    item_finances = ItemFinance.query.filter_by(database=database, item=from_item).first()
    print(to_item_ids)
    if int(from_item_id) in to_item_ids:
        to_item_ids.remove(int(from_item_id))
    for item_id in to_item_ids:
        to_item = Item.query.filter_by(database=database, id=item_id).first()
        if "PRIMARY UNIT" in checkboxes:
            to_item.unit = from_item.unit
            db.session.commit()
        if "JOB RATE" in checkboxes:
            to_item.rate = from_item.rate
            db.session.commit()
        if "BOM_FLAG" in checkboxes:
            to_item.raw_flag = from_item.raw_flag
            db.session.commit()
        if "BOM" in checkboxes:
            item_parent_tree=[to_item]
            search_space = [to_item]
            while len(search_space) > 0:
                item = search_space[0]
                search_space.remove(item)
                parent_items_bom = BOM.query.filter_by(database=database, child_item=item).all()
                for parent_item_bom in parent_items_bom:
                    if parent_item_bom.parent_item.id == to_item.id:
                        db.session.delete(parent_item_bom)
                    item_parent_tree.append(parent_item_bom.parent_item)
                    search_space.append(parent_item_bom.parent_item)
            # to_item.raw_flag = "NO"
            to_item_bom = BOM.query.filter_by(database=database, parent_item=to_item).all()
            for to_item_bom_i in to_item_bom:
                db.session.delete(to_item_bom_i) 
            for bom_item in bom_items:
                if bom_item.child_item in item_parent_tree:
                    # res+= ""
                    continue
                to_bom_item = BOM(parent_item=to_item, child_item=bom_item.child_item, child_item_qty=bom_item.child_item_qty, child_item_unit=bom_item.child_item_unit, 
                    database=database)
                db.session.add(to_bom_item)
                db.session.commit()
        if "UNIT RELATIONS" in checkboxes:
            to_unit_relations = ItemUnit.query.filter_by(database=database, item=to_item).all()
            for relation in item_units:
                to_unit_relation = ItemUnit(database=database, item=to_item, unit_name=relation.unit_name, conversion_factor=relation.conversion_factor,
                    unit_type = relation.unit_type)
                db.session.add(to_unit_relation)
                db.session.commit()
            for to_unit_relations_i in to_unit_relations:
                db.session.delete(to_unit_relations_i) 
        if "CATEGORIES" in checkboxes:
            to_unit_categories= ItemCategory.query.filter_by(database=database, item=to_item).all()
            for item_category in item_categories:
                to_item_category = ItemCategory(database=database, item=to_item, category = item_category.category)
                db.session.add(to_item_category)
                db.session.commit()
            for to_unit_categories_i in to_unit_categories:
                db.session.delete(to_unit_categories_i) 
        if "INVENTORY" in checkboxes:
            to_item_inventory = to_item.iteminventory
            if not to_item_inventory:
                item_inv = ItemInventory(database=database, item=to_item)
                db.session.add(item_inv)
                db.session.commit()
                to_item_inventory = item_inv
            from_item_inventory = from_item.iteminventory
            if not from_item_inventory:
                item_inv = ItemInventory(database=database, item=from_item)
                db.session.add(item_inv)
                db.session.commit()
                from_item_inventory = item_inv
            to_item_inventory.consumption_mode = from_item_inventory.consumption_mode
            to_item_inventory.min_level = from_item_inventory.min_level
            to_item_inventory.max_level = from_item_inventory.max_level
            db.session.commit()
        if "FINANCE" in checkboxes:
            to_item_finance = to_item.itemfinance
            if not to_item_finance:
                item_fin = ItemFinance(database=database, item=to_item)
                db.session.add(item_fin)
                db.session.commit()
                to_item_finance = item_fin
            from_item_finance = from_item.itemfinance
            if not from_item_finance:
                item_fin = ItemFinance(database=database, item=from_item)
                db.session.add(item_fin)
                db.session.commit()
                from_item_finance = item_fin
            to_item_finance.hsn_code = from_item_finance.hsn_code
            to_item_finance.sale_price = from_item_finance.sale_price
            to_item_finance.cost_price = from_item_finance.cost_price
            to_item_finance.tax = from_item_finance.tax
            db.session.commit()
        res+= f"{to_item.name},"
    res+= ". Copied Successfully"
    return res